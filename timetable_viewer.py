# Logic
import os
import csv
from typing import Optional
from datetime import datetime, timedelta
from dateutil import parser

# Excel Export
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter

# PDF Export
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, A2
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle

# GUI
import tkinter
import customtkinter
from tkinter import ttk, Listbox, messagebox
from tkcalendar import Calendar

# GUI appearance & theme
customtkinter.set_appearance_mode("System")
customtkinter.set_default_color_theme("blue")

class Schedule:
    def __init__(self, name, description, date, day, start_time, end_time, duration, location, size, lecturer, zone):
        name = name.split("_")
        self.__cohort:str = "_".join(name[0:2])
        self.__study_mode:str = name[2]
        self.__module_code:str = name[3]
        self.__class_type:str = name[4]
        self.__description:str = description
        self.__date:datetime = datetime.strptime(date, '%d/%m/%Y')
        self.__day:int = parser.parse(day).weekday()
        self.__start_time:datetime = datetime.strptime(start_time, '%H:%M:%S')
        self.__end_time:datetime = datetime.strptime(end_time, '%H:%M:%S')
        self.__duration:datetime = datetime.strptime(duration, '%H:%M')
        self.__lecturer:str = lecturer
        self.__location:str = location
        self.__size:int = int(size)
        self.__zone:str = zone

    def getItem(self, variable):
        match variable:
            case "Cohort":
                return self.__cohort
            case "Study_Mode":
                return self.__study_mode
            case "Module_Code":
                return self.__module_code
            case "Class_Type":
                return self.__class_type
            case "Description":
                return self.__description.split(" (")[0]
            case "Activity date" | "Date" | "Start_Date" | "End_Date":
                return self.__date
            case "Date_str":
                return self.__date.strftime('%d/%m/%Y')
            case "Scheduled Day" | "Day":
                return self.__day
            case "Day_str":
                weekday = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
                return weekday[self.__day]
            case "Scheduled Start Time" | "Start_Time":
                return self.__start_time
            case "Start_Time_str":
                return self.__start_time.strftime('%H:%M:%S')
            case "Scheduled End Time" | "End_Time":
                return self.__end_time
            case "End_Time_str":
                return self.__end_time.strftime('%H:%M:%S')
            case "Time":
                return f"{self.__start_time.strftime('%H:%M')} ~ {self.__end_time.strftime('%H:%M')}"
            case "Duration":
                return self.__duration.strftime('%H:%M')
            case "Allocated Staff Name" | "Lecturer":
                return self.__lecturer
            case "Allocated Location Name" | "Location":
                return self.__location
            case "Planned Size" | "Size" | "Min_Size" | "Max_Size":
                return self.__size
            case "Size_str":
                return str(self.__size)
            case "Zone Name" | "Zone":
                return self.__zone
            case "Date_Time":
                return datetime(self.__date.year, self.__date.month, self.__date.day, self.__start_time.hour, self.__start_time.minute, self.__start_time.second)
            case _:
                raise Exception("Wrong input type")

class File:
    def __init__(self, directoryPath, fileName):
        self.__fileName:str = fileName
        self.__directoryPath:str = directoryPath
        self.__validity:bool = self.__validateFormat()

    def __validateFormat(self):
        with open(self.getPath()) as csv_file:
            dict_reader = csv.DictReader(csv_file)
            headers = dict_reader.fieldnames
            rows = 0
    
            for row in dict_reader:
                for value in row.values():
                    if value == None:
                        return False
                rows += 1
            
            # file ends with .csv
            # header 12 len
            # row at least 1
            # each row has 12 field
            if not self.__fileName.endswith(".csv") or len(headers)!=12 or rows<1: # type: ignore
                return False

        return True

    def getPath(self):
        return self.__directoryPath + self.__fileName
    
    def getFileName(self):
        return self.__fileName

    def getValidity(self):
        return self.__validity

class ScheduleHandler:
    def __init__(self):
        self.__files:list[File] = []
        self.__schedules:list[Schedule] = []
        self.__ignoreFiles:list[str] = []

    def resetHandler(self):
        self.__files:list[File] = []
        self.__schedules:list[Schedule] = []
    
    def loadDirectory(self, directoryPath):
        for file in os.listdir(directoryPath):
            if file.endswith(".csv") and file not in self.__ignoreFiles:
                new_file = File(directoryPath, file)
                self.__files.append(new_file)

        # load schedules
        for file in self.__files:

            if file.getValidity() == True and file.getFileName() not in self.__ignoreFiles:
                with open(file.getPath()) as csv_file:
                    csv_reader = csv.reader(csv_file)
                    # skip the first row
                    next(csv_reader)
                    for schedule in csv_reader:
                        new_schedule = Schedule(schedule[1], schedule[2], schedule[3], schedule[4], schedule[5], schedule[6], schedule[7], schedule[8], schedule[9], schedule[10], schedule[11])
                        self.__schedules.append(new_schedule)

    def getFiles(self):
        return self.__files
    
    def getSchedules(self):
        return self.__schedules

    def addIgnore(self, fileName):
        self.__ignoreFiles.append(fileName)

# Data Structure & Sorting Algorithm
class Heap:
    def __init__(self, schedules, variable):
        self.__unsortedSchedule:list[Schedule] = schedules
        self.__heap:list[Schedule] = []
        self.__variable:Optional[str] = variable

    def __getParent(self, index):
        return (index-1)//2

    def __getLeftChild(self, index):
        return (index*2)+1

    def __getRightChild(self, index):
        return (index*2)+2

    def __swap(self, index, parent_index):
        self.__heap[index], self.__heap[parent_index] = self.__heap[parent_index], self.__heap[index]

    def __insert(self, schedule):
        # insert at last index
        self.__heap.append(schedule)
        index = len(self.__heap) - 1
        # check parent node and swap accordingly
        while index >= 0:
            parent_index = self.__getParent(index)
            if parent_index >= 0 and self.__heap[parent_index].getItem(self.__variable) > self.__heap[index].getItem(self.__variable):  # type: ignore
                self.__swap(index, parent_index)
                index = parent_index
            else:
                break

    def __popMin(self):
        last_index = len(self.__heap) - 1
        if last_index < 0:
            return False
        self.__swap(0, last_index)
        min = self.__heap.pop()
        self.__minHeapify(0)
        return min

    def __minHeapify(self, index):
        left_index = self.__getLeftChild(index)
        right_index = self.__getRightChild(index)
        min_index = index

        if left_index <= len(self.__heap) - 1 and self.__heap[min_index].getItem(self.__variable) > self.__heap[left_index].getItem(self.__variable):
            min_index = left_index
        if right_index <= len(self.__heap) - 1 and self.__heap[min_index].getItem(self.__variable) > self.__heap[right_index].getItem(self.__variable):
            min_index = right_index
        
        if min_index != index:
            self.__swap(index, min_index)
            self.__minHeapify(min_index)

    def heapify(self):
        for schedule in self.__unsortedSchedule:
            self.__insert(schedule)

    def listify(self):
        sortedSchedules = []
        while self.__heap != []:
            sortedSchedules.append(self.__popMin())
        return sortedSchedules

class Sorter:
    def sort(self, schedules, variable, descending=False):
        heap = Heap(schedules, variable)
        heap.heapify()
        sortedSchedules = heap.listify()
        if descending:
            sortedSchedules.reverse()
        return sortedSchedules

class Filter(Sorter):
    def filter(self, schedules, variable, value):
        sortedSchedules = self.sort(schedules, variable)
        filteredSchedules = []
        match variable:
            case "Date" | "Start_Time" | "End_Time" | "Size":
                filteredSchedules.extend(self.__binaryRangeSearch(sortedSchedules, variable, value=value))
            case  "Start_Date":
                filteredSchedules.extend(self.__binaryRangeSearch(sortedSchedules, variable, min_value=value))
            case "End_Date":
                filteredSchedules.extend(self.__binaryRangeSearch(sortedSchedules, variable, max_value=value))
            case _:
                values = value.split("&&&")
                for value in values:
                    filteredSchedules.extend(self.__binaryRangeSearch(sortedSchedules, variable, value=value))
        return filteredSchedules
 
    def __binaryRangeSearch(self, schedules, variable, value=None, min_value=None, max_value=None):
            # binary search
            start = 0
            end = len(schedules)-1
            if value != None:
                while start<=end:
                    mid = (start+end)//2
                    if schedules[mid].getItem(variable) == value:
                        break
                    elif schedules[mid].getItem(variable) < value:
                        start = mid+1
                    else:
                        end = mid-1
                else:
                    return []
            
                # get range index
                current_min = mid
                current_max = mid
                next_min = mid - 1
                next_max = mid + 1

                while next_min>=0 and schedules[next_min].getItem(variable) == value:
                    current_min = next_min
                    next_min -= 1
                    
                while next_max<=len(schedules)-1 and schedules[next_max].getItem(variable) == value:
                    current_max = next_max
                    next_max += 1
                
                return schedules[current_min:current_max+1]
            
            elif min_value != None:
                while start<=end:
                    mid = (start+end)//2
                    if schedules[mid].getItem(variable) >= min_value: # type: ignore
                        if mid-1 < 0:
                            return schedules
                        elif schedules[mid-1].getItem(variable) < min_value: # type: ignore
                            return schedules[mid:]
                        end = mid-1
                    else:
                        start = mid+1
                else:
                    return []
            
            elif max_value != None:
                while start<=end:
                    mid = (start+end)//2
                    if schedules[mid].getItem(variable) <= max_value:
                        if mid+1>len(schedules)-1:
                            return schedules
                        elif schedules[mid+1].getItem(variable) > max_value:
                            return schedules[:mid+1]
                        start = mid+1
                    else:
                        end = mid-1
                        
                else:
                    return []
            
            else:
                return []
                
class ScheduleController(Filter):
    def __init__(self, handler):
        self.__processedSchedules:list[Schedule] = handler.getSchedules()

    def control(self, sortBy, **query):
        # filter
        for variable, value in query.items():
            if value != None:
                self.__processedSchedules = self.filter(self.__processedSchedules, variable, value)

        # sort
        self.sortProcessed(sortBy)

    def sortProcessed(self, sortBy, descending=False):
        self.__processedSchedules = self.sort(self.__processedSchedules, sortBy)
        if descending == True:
            self.__processedSchedules.reverse()

    def getProcessed(self):
        return self.__processedSchedules

    def getValuesSet(self, variable):
        itemsSet = []
        for schedule in self.sort(self.__processedSchedules, variable):
            value = schedule.getItem(variable)
            if value not in itemsSet:
                itemsSet.append(value)
        return itemsSet
    
    def getModuleSet(self):
        itemsSet = []
        moduleSet = []
        for schedule in self.sort(self.__processedSchedules, "Description"):
            value = schedule.getItem("Description")
            if value not in itemsSet:
                itemsSet.append(value)
                moduleSet.append(schedule)
        return moduleSet
    
    def getItems(self, variable):
        items = []
        for schedule in self.sort(self.__processedSchedules, variable):
            value = schedule.getItem(variable)
            items.append(value)
        return items
    
    def getMaxDuplicate(self):
        items = self.getItems("Date")
        new_list = {}
        max = 0
        for item in items:
            if item not in new_list:
                new_list[item] = 0
            new_list[item] += 1
        for item in new_list:
            if new_list[item] > max:
                max = new_list[item]
        return max

class TimetableBuilder:
    def __init__(self, controller):
        self.controller:ScheduleController = controller

    def export(self, format, name, path):
        if format == "xlsx":
            self.workbook = Workbook()
            self.worksheet = self.workbook.active
            self.duplicates = self.controller.getMaxDuplicate()
            self.__setStyle()
            self.__getDateRange()

            self.__setHeight()
            self.__buildModule()
            self.__buildCalendar(self.last_row+3, start_column=1)
            self.__insert_cell()

            self.workbook.save(f"{path}/{name}.{format}")
            
        elif format == "pdf":
            self.controller.sortProcessed("Date_Time")
            schedules = self.controller.getProcessed()
            data = [["No","Cohort","Study Mode", "Lecturer", "Module Code", "Description", "Date", "Day", "Start Time", "End Time", "Duration", "Class Type", "Location", "Size", "Zone"]]
            for id, schedule in enumerate(schedules):
                row = [f"{id+1}", f"{schedule.getItem('Cohort')}", f"{schedule.getItem('Study_Mode')}", f"{schedule.getItem('Lecturer')}", f"{schedule.getItem('Module_Code')}", f"{schedule.getItem('Description')}", f"{schedule.getItem('Date_str')}", f"{schedule.getItem('Day_str')}", f"{schedule.getItem('Start_Time_str')}", f"{schedule.getItem('End_Time_str')}", f"{schedule.getItem('Duration')}", f"{schedule.getItem('Class_Type')}", f"{schedule.getItem('Location')}", f"{schedule.getItem('Size')}", f"{schedule.getItem('Zone')}"]
                data.append(row)
            pdf_path = f"{path}/{name}.{format}"
            pdf = SimpleDocTemplate(pdf_path, pagesize=landscape(A2))
            table = Table(data)
            style = TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.skyblue), ('ALIGN', (0, 0), (-1, -1), 'CENTER'), ('GRID', (0, 0), (-1, -1), 1, colors.black)])
            table.setStyle(style)
            pdf.build([table])

    # calculate
    def __getPosition(self, row, column, position):
        match position:
            case 0 | 1 | 2:
                cell_row = row + 1
            case 3 | 4 | 5:
                cell_row = row + 5
            case 6 | 7 | 8:
                cell_row = row + 9
            case 9 | 10 | 11:
                cell_row = row + 13

        match position:
            case 0 | 3 | 6 | 9:
                cell_column = column
            case 1 | 4 | 7 | 10:
                cell_column = column + 1
            case 2 | 5 | 8 | 11:
                cell_column = column + 2

        return cell_row, cell_column

    def __calculatePosition(self, start_row, start_column, weeks):
        self.weeks_rows = [start_row+((self.height+1)*i)+1 for i in range(weeks)]
        self.days_columns = [start_column+(3*i)+1 for i in range(7)]

    def __setHeight(self):
        if self.duplicates >= 1:
            self.height = 4
            self.blocks = 3
        if self.duplicates >= 4:
            self.height = 8
            self.blocks = 6
        if self.duplicates >= 7:
            self.height = 12
            self.blocks = 9
        if self.duplicates >= 10:
            self.height = 16
            self.blocks = 12

    def __getDateRange(self):
        self.controller.sortProcessed("Date_Time")
        self.schedules = self.controller.getProcessed()
        self.start_date = self.schedules[0].getItem("Date")
        self.end_date = self.schedules[-1].getItem("Date")
        first_week = (self.start_date - timedelta(days=self.start_date.weekday()))
        last_week = (self.end_date - timedelta(days=self.end_date.weekday()))
        self.total_weeks = (last_week - first_week).days//7 +1

    def __getDate(self, row_index, column_index):
        start_day = self.start_date.weekday()
        weekday = [-6,-5,-4,-3,-2,-1,0,1,2,3,4,5,6]

        match start_day:
            case 0:
                new_weekday = weekday[6:]
            case 1:
                new_weekday = weekday[5:-1]
            case 2:
                new_weekday = weekday[4:-2]
            case 3:
                new_weekday = weekday[3:-3]
            case 4:
                new_weekday = weekday[2:-4]
            case 5:
                new_weekday = weekday[1:-5]
            case 6:
                new_weekday = weekday[0:-6]

        total_days = new_weekday[column_index] + (7 * row_index)

        result_date = self.start_date + timedelta(days=total_days)
        return result_date.strftime('%d/%m/%Y')

    # build
    def __applyCellBorder(self):
        for row in self.weeks_rows: # row
            for column in self.days_columns: # column
                for position in range(self.blocks): # 0 - 16
                    position_row, position_column = self.__getPosition(row, column, position)
                    self.__setCellBorder(position, position_row, position_column)
    
    def __setCellBorder(self, position, start_row, column):
        cell1 = self.worksheet.cell(row=start_row, column=column)
        cell2 = self.worksheet.cell(row=start_row+1, column=column)
        cell3 = self.worksheet.cell(row=start_row+2, column=column)
        cell4 = self.worksheet.cell(row=start_row+3, column=column)

        match position:
            case 0 | 3 | 6 | 9:
                cell1.border = Border(left=self.thick, right=self.thin)
                cell2.border = Border(left=self.thick, right=self.thin)
                cell3.border = Border(left=self.thick, right=self.thin)
                cell4.border = Border(bottom=self.thin, left=self.thick, right=self.thin)
            case 1 | 4 | 7 | 10:
                cell1.border = Border(left=self.thin, right=self.thin)
                cell2.border = Border(left=self.thin, right=self.thin)
                cell3.border = Border(left=self.thin, right=self.thin)
                cell4.border = Border(bottom=self.thin, left=self.thin, right=self.thin)
            case 2 | 5 | 8 | 11:
                cell1.border = Border(left=self.thin, right=self.thick)
                cell2.border = Border(left=self.thin, right=self.thick)
                cell3.border = Border(left=self.thin, right=self.thick)
                cell4.border = Border(bottom=self.thin, left=self.thin, right=self.thick)
            
        cell1.fill = self.white
        cell2.fill = self.white
        cell3.fill = self.white
        cell4.fill = self.white

    def __buildModule(self):
        current_row = 3
        self.color_set = {}
        for schedule in self.controller.getModuleSet():
            if schedule.getItem("Description") not in self.color_set.keys():
                self.color_set[schedule.getItem("Description")] = self.color_patterns.pop(0)

            cohort_cell = self.worksheet.cell(row=current_row+1, column=13)
            code_cell = self.worksheet.cell(row=current_row+1, column=14)
            module_cell = self.worksheet.cell(row=current_row+1, column=15)
            lecturer_cell = self.worksheet.cell(row=current_row+1, column=19)

            cohort_cell.value = schedule.getItem("Cohort")
            code_cell.value = schedule.getItem("Module_Code")
            module_cell.value = schedule.getItem("Description")
            lecturer_cell.value = schedule.getItem("Lecturer")

            color = self.color_set[schedule.getItem("Description")]
            code_cell.fill = color


            current_row += 1
        self.last_row = current_row
        if self.last_row < 7:
            self.last_row = 7
            
    def __buildCalendar(self, start_row, start_column):
        self.__calculatePosition(start_row, start_column, self.total_weeks)
        self.__applyCellBorder()
        
        week_cell = self.worksheet.cell(row=self.weeks_rows[0]-1, column=start_column)
        week_cell.value = "Week"
        week_cell.alignment = Alignment(horizontal='center')
        week_cell.border = self.fullBorder
        week_cell.fill = self.blue

        self.worksheet.merge_cells(start_row=3, start_column=2, end_row=7, end_column=7)
        title_cell = self.worksheet.cell(row=3, column=2)
        title_cell.value = "PSB Timetable"
        title_cell.font = self.titleFont

        cohort_cell = self.worksheet.cell(row=3, column=13)
        cohort_cell.value = "Cohort"
        cohort_cell.font = self.headerFont
        code_cell = self.worksheet.cell(row=3, column=14)
        code_cell.value = "Code"
        code_cell.font = self.headerFont
        module_cell = self.worksheet.cell(row=3, column=15)
        module_cell.value = "Module"
        module_cell.font = self.headerFont
        lecturer_cell = self.worksheet.cell(row=3, column=19)
        lecturer_cell.value = "Value"
        lecturer_cell.font = self.headerFont

        weekday = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
        for index, day_column in enumerate(self.days_columns): # Weekday
            self.worksheet.merge_cells(start_row=start_row, start_column=day_column, end_row=start_row, end_column=day_column+2)
            cell1 = self.worksheet.cell(row=start_row, column=day_column)
            cell2 = self.worksheet.cell(row=start_row, column=day_column+1)
            cell3 = self.worksheet.cell(row=start_row, column=day_column+2)
            cell1.border = self.fullBorder
            cell2.border = self.fullBorder
            cell3.border = self.fullBorder
            cell1.fill = self.blue
            cell2.fill = self.blue
            cell3.fill = self.blue

            
            cell1.value = weekday[index]
            cell1.alignment = Alignment(horizontal='center', vertical='center')

        for row_index, week_row in enumerate(self.weeks_rows): # Weeks
            for column_index, day_column in enumerate(self.days_columns): # Date
                self.worksheet.merge_cells(start_row=week_row, start_column=day_column, end_row=week_row, end_column=day_column+2)
                cell1 = self.worksheet.cell(row=week_row, column=day_column)
                cell2 = self.worksheet.cell(row=week_row, column=day_column+1)
                cell3 = self.worksheet.cell(row=week_row, column=day_column+2)
                cell1.border = self.fullBorder
                cell2.border = self.fullBorder
                cell3.border = self.fullBorder
                cell1.alignment = Alignment(horizontal='center', vertical='center')
                cell1.fill = self.grey
                date = self.__getDate(row_index, column_index)
                cell1.value = date

                # add dates here
            
            self.worksheet.merge_cells(start_row=week_row, start_column=start_column, end_row=week_row+self.height, end_column=start_column)
            cell0 = self.worksheet.cell(row=week_row, column=start_column)
            cell0.value = row_index+1
            cell0.alignment = Alignment(horizontal='center', vertical='center')
            cell0.border = self.fullBorder
            cell0.fill = self.blue
            if self.height >= 4:
                cell1 = self.worksheet.cell(row=week_row+1, column=start_column)
                cell2 = self.worksheet.cell(row=week_row+2, column=start_column)
                cell3 = self.worksheet.cell(row=week_row+3, column=start_column)
                cell4 = self.worksheet.cell(row=week_row+4, column=start_column)
                cell1.border = self.fullBorder
                cell2.border = self.fullBorder
                cell3.border = self.fullBorder
                cell4.border = self.fullBorder
            if self.height >= 8:
                cell5 = self.worksheet.cell(row=week_row+5, column=start_column)
                cell6 = self.worksheet.cell(row=week_row+6, column=start_column)
                cell7 = self.worksheet.cell(row=week_row+7, column=start_column)
                cell8 = self.worksheet.cell(row=week_row+8, column=start_column)
                cell5.border = self.fullBorder
                cell6.border = self.fullBorder
                cell7.border = self.fullBorder
                cell8.border = self.fullBorder
            if self.height >= 12:
                cell9 = self.worksheet.cell(row=week_row+9, column=start_column)
                cell10 = self.worksheet.cell(row=week_row+10, column=start_column)
                cell11 = self.worksheet.cell(row=week_row+11, column=start_column)
                cell12 = self.worksheet.cell(row=week_row+12, column=start_column)
                cell9.border = self.fullBorder
                cell10.border = self.fullBorder
                cell11.border = self.fullBorder
                cell12.border = self.fullBorder
            if self.height >= 16:
                cell13 = self.worksheet.cell(row=week_row+13, column=start_column)
                cell14 = self.worksheet.cell(row=week_row+14, column=start_column)
                cell15 = self.worksheet.cell(row=week_row+15, column=start_column)
                cell16 = self.worksheet.cell(row=week_row+16, column=start_column)
                cell13.border = self.fullBorder
                cell14.border = self.fullBorder
                cell15.border = self.fullBorder
                cell16.border = self.fullBorder

        start_column + 7
        for i in range(start_column+1, start_column+(7*3)+1):
            self.worksheet.column_dimensions[get_column_letter(i)].width = 15

    def __insert_cell(self):
        schedules = self.controller.getProcessed()
        for row_index, row in enumerate(self.weeks_rows):
            for column_index, column in enumerate(self.days_columns):
                current_date = self.__getDate(row_index,column_index)
                for cell in range(self.blocks):
                    cell_row, cell_column = self.__getPosition(row, column, cell)
                    if len(schedules) == 0:
                        break
                    elif schedules[0].getItem("Date_str") == current_date:
                        schedule = schedules.pop(0)
                        color = self.color_set[schedule.getItem("Description")]

                        code_cell = self.worksheet.cell(row=cell_row, column=cell_column)
                        time_cell = self.worksheet.cell(row=cell_row+1, column=cell_column)
                        lec_cell = self.worksheet.cell(row=cell_row+2, column=cell_column)
                        location_cell = self.worksheet.cell(row=cell_row+3, column=cell_column)

                        code_value = schedule.getItem("Module_Code")
                        time_value = schedule.getItem("Time")
                        lec_value = schedule.getItem("Class_Type")
                        location_value = schedule.getItem("Location")

                        code_cell.value = code_value
                        time_cell.value = time_value
                        lec_cell.value = lec_value
                        location_cell.value = location_value

                        code_cell.alignment = Alignment(horizontal='center')
                        time_cell.alignment = Alignment(horizontal='center')
                        lec_cell.alignment = Alignment(horizontal='center')
                        location_cell.alignment = Alignment(horizontal='center')

                        code_cell.fill = color
                        time_cell.fill = color
                        lec_cell.fill = color
                        location_cell.fill = color

    # utility
    def __setStyle(self):
        self.white = PatternFill(start_color="ffffff", end_color="ffffff", fill_type="solid")
        color_list = ["f85751", "fc8058", "f6a667", "ffc58f", "fffea4", "d0f3a7", "b1d89d", "6bf2fd", "81c5f2", "a39de1", "d0b8e9", "f7c4cd", "f5e0e9"]
        self.color_patterns = [PatternFill(start_color=color, end_color=color, fill_type="solid") for color in color_list]
        self.thin = Side(style='thin')
        self.thick = Side(style='medium')
        self.fullBorder = Border(top=self.thick, bottom=self.thick, left=self.thick, right=self.thick)
        self.blue = PatternFill(start_color="c2dfff", end_color="c2dfff", fill_type="solid")
        self.grey = PatternFill(start_color="e0e0e0", end_color="e0e0e0", fill_type="solid")

        self.titleFont = Font(size = "65", bold=True)
        self.headerFont = Font(size = "12", bold=True, underline="single")

class App(customtkinter.CTk):
    def __init__(self):
        super().__init__()
        self.handler = ScheduleHandler()
        self.controller = ScheduleController(self.handler)
        self.__reset_click_count()

        # configure window
        self.title("Timetable Viewer")
        self.geometry(f"{1100}x{580}")
        self.resizable(width=True, height=True)
        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(0, weight=0)
        self.grid_columnconfigure(1, weight=1)

        ######################################################################################################################################################

        # sidebar outer frame
        self.F_sidebar_out = customtkinter.CTkFrame(self)
        self.F_sidebar_out.grid(row=0, column=0, sticky='nswe')
        self.F_sidebar_out.grid_rowconfigure(0, weight=1)
        self.F_sidebar_out.grid_columnconfigure(0, weight=1)

            # sidebar inner frame
        self.F_sidebar = customtkinter.CTkFrame(self.F_sidebar_out, corner_radius=0, fg_color="transparent")
        self.F_sidebar.grid(row=0, column=0, sticky='nswe', padx=15, pady=(0,5))
        self.F_sidebar.grid_rowconfigure(0, weight=1)
        self.F_sidebar.grid_rowconfigure(1, weight=15)
        self.F_sidebar.grid_columnconfigure(0, weight=1)

                # sidebar inner top
        self.F_sidebar_top = customtkinter.CTkFrame(self.F_sidebar, corner_radius=0)
        self.F_sidebar_top.grid(row=0, column=0, sticky='nswe')
        self.F_sidebar_top.grid_rowconfigure(0, weight=1)
        self.F_sidebar_top.grid_columnconfigure(0, weight=1)

                    # sidebar inner top (Logo)
        self.logo = customtkinter.CTkLabel(self.F_sidebar_top, text="Timetable\nViewer", font=customtkinter.CTkFont(size=24, weight="bold"))
        self.logo.grid(row=0, column=0, sticky='nswe', pady=(10,0))

        ####################################################################################################

                # sidebar inner bottom
        self.F_sidebar_bottom = customtkinter.CTkFrame(self.F_sidebar, corner_radius=0)
        self.F_sidebar_bottom.grid(row=1, column=0, sticky='nswe')
        self.F_sidebar_bottom.grid_rowconfigure((0,2), weight=0)
        self.F_sidebar_bottom.grid_rowconfigure(1, weight=1)
        self.F_sidebar_bottom.grid_columnconfigure((0,1,2), weight=1)
        self.F_sidebar_bottom.grid_columnconfigure(3, weight=0)

                # sidebar inner bottom (Remove Button)
        self.remove_button = customtkinter.CTkButton(self.F_sidebar_bottom, width=28, text="Remove",font=customtkinter.CTkFont(size=12, weight="bold"), fg_color="transparent", border_width=2, text_color=("gray10", "#DCE4EE"), command=self.removeButtonPressed)
        self.remove_button.grid(row=0, column=0, sticky='w', padx=5, pady=10)

                # sidebar inner bottom (Files Table)
        self.files_table = ttk.Treeview(self.F_sidebar_bottom, columns=('No', "File", "Valid"), show='headings', padding=(10,10,5,5))
        self.files_table.grid(row=1, column=0, sticky='nswe', columnspan=4, padx=5, pady=(0,10))
        self.files_table.heading("No", text="No")
        self.files_table.column("No", minwidth=15, width=15)
        self.files_table.heading("File", text="File")
        self.files_table.column("File", minwidth=110, width=110)
        self.files_table.heading("Valid", text="Valid")
        self.files_table.column("Valid", minwidth=40, width=40)

                # sidebar inner bottom (Directory Path)
        self.entry = customtkinter.CTkEntry(self.F_sidebar_bottom, placeholder_text="C:/download/", width=200)
        self.entry.insert(0, "Set Path")
        self.entry.configure(state="disabled")
        self.entry.grid(row=2, column=0, columnspan=3, padx=(5,0), pady=(0,10), sticky="we")

                # sidebar inner bottom (Import Button)
        self.import_button = customtkinter.CTkButton(self.F_sidebar_bottom, width=28, text="Import", font=customtkinter.CTkFont(size=12, weight="bold"), border_width=2, border_color="grey", text_color=("gray10", "#DCE4EE"), command=self.importButtonPressed)
        self.import_button.grid(row=2, column=3, sticky='we', padx=5, pady=(0,10))

        ######################################################################################################################################################

        # main outer frame
        self.F_main_out = customtkinter.CTkFrame(self)
        self.F_main_out.grid(row=0, column=1, sticky='nswe', padx=15, pady=15)
        self.F_main_out.grid_rowconfigure(0, weight=1)
        self.F_main_out.grid_columnconfigure(0, weight=1)


            # main inner frame
        self.F_main = customtkinter.CTkFrame(self.F_main_out, fg_color="transparent")
        self.F_main.grid(row=0, column=0, sticky='nswe', padx=15, pady=(0,15))
        self.F_main.grid_rowconfigure((0,2,3), weight=0)
        self.F_main.grid_rowconfigure(1, weight=1)
        self.F_main.grid_columnconfigure(0, weight=1)

            # main inner tabbed frame (Filtering Options)
        self.tabview = customtkinter.CTkTabview(self.F_main, height=200, fg_color="#242424")
        self.tabview.grid(row=0, column=0, sticky="nsew")
        self.tabview.add("Module")
        self.tabview.add("Date & Time")
        self.tabview.add("Lecture Room")

        ####################################################################################################

                # Tabbed Frame (Module)
        self.tabview.tab("Module").grid_rowconfigure((0,1,2,3), weight=1)
        self.tabview.tab("Module").grid_columnconfigure((0,1,2,3,5,6,7,8,9,10,11), weight=1)
        self.tabview.tab("Module").grid_columnconfigure(5, weight=0)

        self.label = customtkinter.CTkLabel(self.tabview.tab("Module"), text="Cohort")
        self.label.grid(row=0, column=4, sticky='e')
        self.label = customtkinter.CTkLabel(self.tabview.tab("Module"), text=":")
        self.label.grid(row=0, column=5, padx=5)

        self.__load_cohort_option(init=True)

        self.label = customtkinter.CTkLabel(self.tabview.tab("Module"), text="Study Mode")
        self.label.grid(row=1, column=4, sticky='e')
        self.label = customtkinter.CTkLabel(self.tabview.tab("Module"), text=":")
        self.label.grid(row=1, column=5, padx=5)

        self.__load_study_mode_option(init=True)

        self.label = customtkinter.CTkLabel(self.tabview.tab("Module"), text="Lecturer")
        self.label.grid(row=2, column=4, sticky='e')
        self.label = customtkinter.CTkLabel(self.tabview.tab("Module"), text=":")
        self.label.grid(row=2, column=5, padx=5)

        self.__load_lecturer_option(init=True)

        self.label = customtkinter.CTkLabel(self.tabview.tab("Module"), text="Module Code")
        self.label.grid(row=3, column=4, sticky='e')
        self.label = customtkinter.CTkLabel(self.tabview.tab("Module"), text=":")
        self.label.grid(row=3, column=5, padx=5)

        self.__load_module_code_option(init=True)

        self.label = customtkinter.CTkLabel(self.tabview.tab("Module"), text="Description")
        self.label.grid(row=0, column=7)

        self.__load_description_option(init=True)

        self.label = customtkinter.CTkLabel(self.tabview.tab("Module"), text="")
        self.label.grid(row=0, column=11, sticky='w')

        ##################################################

                # Tabbed Frame (Date & Time)
        self.tabview.tab("Date & Time").grid_rowconfigure((0,1,2,3,4), weight=1)
        self.tabview.tab("Date & Time").grid_columnconfigure((1,2,3,5,6,7,8,9,10), weight=1)
        self.tabview.tab("Date & Time").grid_columnconfigure(5, weight=0)
        self.tabview.tab("Date & Time").grid_columnconfigure((0,11), weight=1)

        self.checked = tkinter.IntVar()
        self.check = customtkinter.CTkCheckBox(self.tabview.tab("Date & Time"), text="By Range",variable=self.checked, onvalue=1, offvalue=0, width=15, height=15, command=self.__boxChecked)
        self.check.grid(row=4, column=0, sticky="w")

        self.date_label = customtkinter.CTkLabel(self.tabview.tab("Date & Time"), text="Date")
        self.date_label.grid(row=0, column=4, pady=(10,0), sticky='e')
        self.date_space_label = customtkinter.CTkLabel(self.tabview.tab("Date & Time"), text=":")
        self.date_space_label.grid(row=0, column=5, padx=5, pady=(10,0))

        self.date_option = customtkinter.CTkEntry(self.tabview.tab("Date & Time"), height=15, width=95) # xxx
        self.date_option.insert(0, "dd/mm/yyyy")
        self.date_option.grid(row=0, column=6, pady=(10,0), sticky="w")
        self.date_option.bind("<1>", self.__date_option_clicked)

        self.label = customtkinter.CTkLabel(self.tabview.tab("Date & Time"), text="Duration")
        self.label.grid(row=2, column=4, sticky='e')
        self.label = customtkinter.CTkLabel(self.tabview.tab("Date & Time"), text=":")
        self.label.grid(row=2, column=5, padx=5)

        self.__load_duration_option(init=True)

        self.label = customtkinter.CTkLabel(self.tabview.tab("Date & Time"), text="Start Time")
        self.label.grid(row=3, column=4, sticky='e')

        self.label = customtkinter.CTkLabel(self.tabview.tab("Date & Time"), text="End Time")
        self.label.grid(row=3, column=6, sticky='w')

        self.__load_start_time_option(init=True)
        self.label = customtkinter.CTkLabel(self.tabview.tab("Date & Time"), text="~")
        self.label.grid(row=4, column=5, padx=5)
        self.__load_end_time_option(init=True)

        self.label = customtkinter.CTkLabel(self.tabview.tab("Date & Time"), text="Day")
        self.label.grid(row=0, column=7)

        self.__load_day_option(init=True)

        self.label = customtkinter.CTkLabel(self.tabview.tab("Date & Time"), text="                         ")
        self.label.grid(row=0, column=11, sticky='w')
        
        ##################################################

                # Tabbed Frame (Lecture Room)
        self.tabview.tab("Lecture Room").grid_rowconfigure((0,1,2), weight=1)
        self.tabview.tab("Lecture Room").grid_columnconfigure((0,1,2,3,5,6,7,8,9,10,11), weight=1)
        self.tabview.tab("Lecture Room").grid_columnconfigure(6, weight=0)

        self.label = customtkinter.CTkLabel(self.tabview.tab("Lecture Room"), text="Class Type")
        self.label.grid(row=0, column=4)

        self.__load_class_type_option(init=True)

        self.label = customtkinter.CTkLabel(self.tabview.tab("Lecture Room"), text="Location")
        self.label.grid(row=0, column=5, sticky='e') # column = 6
        self.label = customtkinter.CTkLabel(self.tabview.tab("Lecture Room"), text=":")
        self.label.grid(row=0, column=6, padx=5) # column = 7
        
        self.__load_location_option(init=True)

        self.label = customtkinter.CTkLabel(self.tabview.tab("Lecture Room"), text="Size")
        self.label.grid(row=1, column=5, sticky='e') # column = 6
        self.label = customtkinter.CTkLabel(self.tabview.tab("Lecture Room"), text=":")
        self.label.grid(row=1, column=6, padx=5) # column = 7

        self.__load_size_option(init=True)

        self.label = customtkinter.CTkLabel(self.tabview.tab("Lecture Room"), text="Zone")
        self.label.grid(row=2, column=5, sticky='e') # column = 6
        self.label = customtkinter.CTkLabel(self.tabview.tab("Lecture Room"), text=":")
        self.label.grid(row=2, column=6, padx=5) # column = 7

        self.__load_zone_option(init=True)

        ####################################################################################################

                # main inner (Schedule Table)
        self.schedule_table = ttk.Treeview(self.F_main, columns=("No", "Cohort", "Study Mode", "Lecturer", "Module Code", "Description", "Date", "Day", "Start Time", "End Time", "Duration", "Class Type", "Location", "Size", "Zone"), show='headings', padding=(10,10,5,5))

        self.schedule_table.column("No", width=50, minwidth=50)
        self.schedule_table.column("Cohort", width=125, minwidth=125)
        self.schedule_table.column("Study Mode", width=80, minwidth=80)
        self.schedule_table.column("Lecturer", width=250, minwidth=250)
        self.schedule_table.column("Module Code", width=90, minwidth=90)
        self.schedule_table.column("Description", width=350, minwidth=400)

        self.schedule_table.column("Date", width=90, minwidth=90)
        self.schedule_table.column("Day", width=80, minwidth=80)
        self.schedule_table.column("Start Time", width=70, minwidth=70)
        self.schedule_table.column("End Time", width=70, minwidth=70)
        self.schedule_table.column("Duration", width=55, minwidth=55)

        self.schedule_table.column("Class Type", width=120, minwidth=120)
        self.schedule_table.column("Location", width=110, minwidth=110)
        self.schedule_table.column("Size", width=40, minwidth=40)
        self.schedule_table.column("Zone", width=60, minwidth=60)

        
        self.schedule_table.heading("No", text="No")
        self.schedule_table.heading("Cohort", text="Cohort", command=self.__cohort_clicked)
        self.schedule_table.heading("Study Mode", text="Study Mode", command=self.__study_mode_clicked)
        self.schedule_table.heading("Lecturer", text="Lecturer", command=self.__lecturer_clicked)
        self.schedule_table.heading("Module Code", text="Module Code", command=self.__module_code_clicked)
        self.schedule_table.heading("Description", text="Description", command=self.__description_clicked)

        self.schedule_table.heading("Date", text="Date", command=self.__date_clicked)
        self.schedule_table.heading("Day", text="Day", command=self.__day_clicked)
        self.schedule_table.heading("Start Time", text="Start Time", command=self.__start_time_clicked)
        self.schedule_table.heading("End Time", text="End Time", command=self.__end_time_clicked)
        self.schedule_table.heading("Duration", text="Duration", command=self.__duration_clicked)

        self.schedule_table.heading("Class Type", text="Class Type", command=self.__class_type_clicked)
        self.schedule_table.heading("Location", text="Location", command=self.__location_clicked)
        self.schedule_table.heading("Size", text="Size", command=self.__size_clicked)
        self.schedule_table.heading("Zone", text="Zone", command=self.__zone_clicked)

        self.schedule_table.grid(row=1, column=0, sticky='nswe', pady=(10,0))
        
        ##################################################

                # main inner (Scrollbar)
        scrollbar = customtkinter.CTkScrollbar(self.F_main, orientation="horizontal")
        scrollbar.configure(command=self.schedule_table.xview)
        self.schedule_table.configure(xscrollcommand=scrollbar.set)
        scrollbar.grid(row=2, column=0, pady=(0,5), sticky='nwse')

        ##################################################

                # main inner bottom frame
        self.F_main_confirm = customtkinter.CTkFrame(self.F_main)
        self.F_main_confirm.grid(row=3, column=0, sticky='nswe')
        self.F_main_confirm.grid_rowconfigure(0, weight=1)
        self.F_main_confirm.grid_columnconfigure((0,2), weight=0)
        self.F_main_confirm.grid_columnconfigure(1, weight=1)

                # main inner bottom (Export Button)
        self.export_button = customtkinter.CTkButton(self.F_main_confirm, width=28, text="Export",font=customtkinter.CTkFont(size=12, weight="bold"), fg_color="transparent", border_width=2, text_color=("gray10", "#DCE4EE"), command=self.exportButtonPressed)
        self.export_button.grid(row=0, column=0, sticky='sw', padx=(0,5))

        self.reload_button = customtkinter.CTkButton(self.F_main_confirm, width=28, text="Reload",font=customtkinter.CTkFont(size=12, weight="bold"), fg_color="transparent", border_width=2, text_color=("gray10", "#DCE4EE"), command=self.reloadButtonPressed)
        self.reload_button.grid(row=0, column=1, sticky='se', padx=(0,5))

                # main inner bottom (Complete Button)
        self.load_button = customtkinter.CTkButton(self.F_main_confirm, width=28, text="Confirm",font=customtkinter.CTkFont(size=12, weight="bold"), border_width=2, border_color="grey", text_color=("gray10", "#DCE4EE"), command=self.confirmButtonPressed)
        self.load_button.grid(row=0, column=2, sticky='se', padx=(5,5))

        ####################################################################################################

    # utility functions
    def __clearFilesTable(self):
        for row in self.files_table.get_children():
            self.files_table.delete(row)
        self.update()

    def __loadFiles(self):
        for index, file in enumerate(self.handler.getFiles()):
            no = index + 1
            file_name = file.getFileName()
            valid = file.getValidity()
            if valid == True:
                valid = ""
            else:
                valid = "False"

            data = (no, file_name, valid)
            self.files_table.insert(parent="", index=tkinter.END, values=data) 

    def __clearSchedulesTable(self):
        for row in self.schedule_table.get_children():
            self.schedule_table.delete(row)
        self.update()

    def __loadSchedules(self):
        for index, schedule in enumerate(self.controller.getProcessed()):
            no = index + 1
            cohort = schedule.getItem("Cohort")
            study_mode = schedule.getItem("Study_Mode")
            lecturer = schedule.getItem("Lecturer")
            module_code = schedule.getItem("Module_Code")
            description = schedule.getItem("Description")

            date = schedule.getItem("Date_str")
            day = schedule.getItem("Day_str")
            start_time = schedule.getItem("Start_Time_str")
            end_time = schedule.getItem("End_Time_str")
            duration = schedule.getItem("Duration")

            class_type = schedule.getItem("Class_Type")
            location = schedule.getItem("Location")
            size = schedule.getItem("Size")
            zone = schedule.getItem("Zone")

            data = (no, cohort, study_mode, lecturer, module_code, description, date, day, start_time, end_time, duration, class_type, location, size, zone)

            self.schedule_table.insert(parent="", index=tkinter.END, values=data)
        if len(self.controller.getProcessed()) == 0 and len(self.handler.getFiles()) != 0:
            self.errorPopup("No Schedules Found!")


    # load combobox
    def __load_cohort_option(self, init=False):
        if init == False:
            self.cohort_option.destroy()
        self.cohort_set = self.controller.getValuesSet("Cohort")
        self.cohort_set.insert(0, "No Filter")
        self.cohort_option = customtkinter.CTkComboBox(self.tabview.tab("Module"), height=15, values=self.cohort_set)
        self.cohort_option.set("No Filter")
        self.cohort_option.grid(row=0, column=6, sticky='w')

    def __load_study_mode_option(self, init=False):
        if init == False:
            self.study_mode_option.destroy()
        self.study_mode_set = self.controller.getValuesSet("Study_Mode")
        self.study_mode_set.insert(0, "No Filter")
        self.study_mode_option = customtkinter.CTkComboBox(self.tabview.tab("Module"), height=15, values=self.study_mode_set)
        self.study_mode_option.set("No Filter")
        self.study_mode_option.grid(row=1, column=6, sticky='w')

    def __load_lecturer_option(self, init=False):
        if init == False:
            self.lecturer_option.destroy()
        self.lecturer_set = self.controller.getValuesSet("Lecturer")
        self.lecturer_set.insert(0, "No Filter")
        self.lecturer_option = customtkinter.CTkComboBox(self.tabview.tab("Module"), height=15, values=self.lecturer_set)
        self.lecturer_option.set("No Filter")
        self.lecturer_option.grid(row=2, column=6, sticky='w')

    def __load_module_code_option(self, init=False):
        if init == False:
            self.module_code_option.destroy()
        self.module_code_set = self.controller.getValuesSet("Module_Code")
        self.module_code_set.insert(0, "No Filter")
        self.module_code_option = customtkinter.CTkComboBox(self.tabview.tab("Module"), height=15, values=self.module_code_set)
        self.module_code_option.set("No Filter")
        self.module_code_option.grid(row=3, column=6, sticky='w')

    def __load_duration_option(self, init=False):
        if init == False:
            self.duration_option.destroy()
        self.duration_set = self.controller.getValuesSet("Duration")
        self.duration_set.insert(0, "No Filter")
        self.duration_option = customtkinter.CTkComboBox(self.tabview.tab("Date & Time"), height=15, width= 95, values=self.duration_set)
        self.duration_option.set("No Filter")
        self.duration_option.grid(row=2, column=6, pady=(10,0), sticky='w')

    def __load_location_option(self, init=False):
        if init == False:
            self.location_option.destroy()
        self.location_set = self.controller.getValuesSet("Location")
        self.location_set.insert(0, "No Filter")
        self.location_option = customtkinter.CTkComboBox(self.tabview.tab("Lecture Room"), height=15, values=self.location_set)
        self.location_option.set("No Filter")
        self.location_option.grid(row=0, column=7, sticky='w')

    def __load_size_option(self, init=False):
        if init == False:
            self.size_option.destroy()
        self.size_set = self.controller.getValuesSet("Size")
        self.size_set = [str(x) for x in self.size_set]
        self.size_set.insert(0, "No Filter")
        self.size_option = customtkinter.CTkComboBox(self.tabview.tab("Lecture Room"), height=15, values=self.size_set)
        self.size_option.set("No Filter")
        self.size_option.grid(row=1, column=7, sticky='w')

    def __load_zone_option(self, init=False):
        if init == False:
            self.zone_option.destroy()

        self.zone_set = self.controller.getValuesSet("Zone")
        self.zone_set.insert(0, "No Filter")
        self.zone_option = customtkinter.CTkComboBox(self.tabview.tab("Lecture Room"), height=15, values=self.zone_set)
        self.zone_option.set("No Filter")
        self.zone_option.grid(row=2, column=7, sticky='w')

    def __load_start_time_option(self, init=False):
        if init == False:
            self.start_time_option.destroy()

        self.start_time_set = self.controller.getValuesSet("Start_Time_str")
        self.start_time_set.insert(0, "No Filter")
        self.start_time_option = customtkinter.CTkComboBox(self.tabview.tab("Date & Time"), height=15, width=100, values=self.start_time_set)
        self.start_time_option.set("No Filter")
        self.start_time_option.grid(row=4, column=4, sticky='e')

    def __load_end_time_option(self, init=False):
        if init == False:
            self.end_time_option.destroy()

        self.end_time_set = self.controller.getValuesSet("End_Time_str")
        self.end_time_set.insert(0, "No Filter")
        self.end_time_option = customtkinter.CTkComboBox(self.tabview.tab("Date & Time"), height=15, width=100, values=self.end_time_set)
        self.end_time_option.set("No Filter")
        self.end_time_option.grid(row=4, column=6, sticky='w')

    # load listbox
    def __load_day_option(self, init=False): 
        if init == False:
            self.day_option.destroy()
        self.day_set = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
        self.day_option = Listbox(self.tabview.tab("Date & Time"), selectmode="multiple", height=2, width=15)
        for value in self.day_set:
            self.day_option.insert(tkinter.END, value)
        self.day_option.grid(row=1, column=7, rowspan=7, pady=(0,5), sticky='ns')

    def __load_class_type_option(self, init=False):
        if init == False:
            self.class_type_option.destroy()
        
        self.class_type_set = self.controller.getValuesSet("Class_Type")
        self.class_type_option = Listbox(self.tabview.tab("Lecture Room"), selectmode="multiple", height=2)
        for value in self.class_type_set:
            self.class_type_option.insert(tkinter.END, value)
        self.class_type_option.grid(row=1, column=4, rowspan=2, pady=(0,5), sticky='nswe')

    def __load_description_option(self, init=False):
        if init == False:
            self.description_option.destroy()
        
        self.description_set = self.controller.getValuesSet("Description")
        self.description_option = Listbox(self.tabview.tab("Module"), selectmode="multiple", height=2, width=30)
        for value in self.description_set:
            self.description_option.insert(tkinter.END, value)
        self.description_option.grid(row=1, column=7, rowspan=3, padx=5, pady=(0,5), sticky='nswe')

    # reload filters
    def __reload_filters(self):
        self.__load_cohort_option()
        self.__load_study_mode_option()
        self.__load_lecturer_option()
        self.__load_module_code_option()
        self.__load_description_option()

        self.__load_duration_option()
        self.__load_start_time_option()
        self.__load_end_time_option()
        self.__load_day_option()

        self.__load_class_type_option()
        self.__load_location_option()
        self.__load_size_option()
        self.__load_zone_option()

        if (self.checked.get() == 1):
            self.start_date_option.delete(0, tkinter.END)
            self.start_date_option.insert(0, "dd/mm/yyyy")
            self.end_date_option.delete(0, tkinter.END)
            self.end_date_option.insert(0, "dd/mm/yyyy")
        else:
            self.date_option.delete(0, tkinter.END)
            self.date_option.insert(0, "dd/mm/yyyy")

    # header sort
    def __reset_click_count(self):
        self.cohort_click_count = 0
        self.study_mode_click_count = 0
        self.lecturer_click_count = 0
        self.module_code_click_count = 0
        self.description_click_count = 0
        self.date_click_count = 0
        self.day_click_count = 0
        self.start_time_click_count = 0
        self.end_time_click_count = 0
        self.duration_click_count = 0
        self.class_type_click_count = 0
        self.location_click_count = 0
        self.size_click_count = 0
        self.zone_click_count = 0

    def __cohort_clicked(self):
        self.__clearSchedulesTable()
        if self.cohort_click_count == 0: #
            self.controller.sortProcessed("Cohort") #
            self.__reset_click_count()
            self.cohort_click_count += 1 #
        elif self.cohort_click_count == 1: #
            self.controller.sortProcessed("Cohort", descending=True) #
            self.__reset_click_count()
        self.__loadSchedules()
        self.update()

    def __study_mode_clicked(self):
        self.__clearSchedulesTable()
        if self.study_mode_click_count == 0: #
            self.controller.sortProcessed("Study_Mode") #
            self.__reset_click_count()
            self.study_mode_click_count += 1 #
        elif self.study_mode_click_count == 1: #
            self.controller.sortProcessed("Study_Mode", descending=True) #
            self.__reset_click_count()
        self.__loadSchedules()
        self.update()
    
    def __lecturer_clicked(self):
        self.__clearSchedulesTable()
        if self.lecturer_click_count == 0: #
            self.controller.sortProcessed("Lecturer") #
            self.__reset_click_count()
            self.lecturer_click_count += 1 #
        elif self.lecturer_click_count == 1: #
            self.controller.sortProcessed("Lecturer", descending=True) #
            self.__reset_click_count()
        self.__loadSchedules()
        self.update()
    
    def __module_code_clicked(self):
        self.__clearSchedulesTable()
        if self.module_code_click_count == 0: #
            self.controller.sortProcessed("Module_Code") #
            self.__reset_click_count()
            self.module_code_click_count += 1 #
        elif self.module_code_click_count == 1: #
            self.controller.sortProcessed("Module_Code", descending=True) #
            self.__reset_click_count()
        self.__loadSchedules()
        self.update()
    
    def __description_clicked(self):
        self.__clearSchedulesTable()
        if self.description_click_count == 0: #
            self.controller.sortProcessed("Description") #
            self.__reset_click_count()
            self.description_click_count += 1 #
        elif self.description_click_count == 1: #
            self.controller.sortProcessed("Description", descending=True) #
            self.__reset_click_count()
        self.__loadSchedules()
        self.update()
    
    def __date_clicked(self):
        self.__clearSchedulesTable()
        if self.date_click_count == 0: #
            self.controller.sortProcessed("Date") #
            self.__reset_click_count()
            self.date_click_count += 1 #
        elif self.date_click_count == 1: #
            self.controller.sortProcessed("Date", descending=True) #
            self.__reset_click_count()
        self.__loadSchedules()
        self.update()
    
    def __day_clicked(self):
        self.__clearSchedulesTable()
        if self.day_click_count == 0: #
            self.controller.sortProcessed("Day") #
            self.__reset_click_count()
            self.day_click_count += 1 #
        elif self.day_click_count == 1: #
            self.controller.sortProcessed("Day", descending=True) #
            self.__reset_click_count()
        self.__loadSchedules()
        self.update()
    
    def __start_time_clicked(self):
        self.__clearSchedulesTable()
        if self.start_time_click_count == 0: #
            self.controller.sortProcessed("Start_Time") #
            self.__reset_click_count()
            self.start_time_click_count += 1 #
        elif self.start_time_click_count == 1: #
            self.controller.sortProcessed("Start_Time", descending=True) #
            self.__reset_click_count()
        self.__loadSchedules()
        self.update()
    
    def __end_time_clicked(self):
        self.__clearSchedulesTable()
        if self.end_time_click_count == 0: #
            self.controller.sortProcessed("End_Time") #
            self.__reset_click_count()
            self.end_time_click_count += 1 #
        elif self.end_time_click_count == 1: #
            self.controller.sortProcessed("End_Time", descending=True) #
            self.__reset_click_count()
        self.__loadSchedules()
        self.update()
    
    def __duration_clicked(self):
        self.__clearSchedulesTable()
        if self.duration_click_count == 0: #
            self.controller.sortProcessed("Duration") #
            self.__reset_click_count()
            self.duration_click_count += 1 #
        elif self.duration_click_count == 1: #
            self.controller.sortProcessed("Duration", descending=True) #
            self.__reset_click_count()
        self.__loadSchedules()
        self.update()
    
    def __class_type_clicked(self):
        self.__clearSchedulesTable()
        if self.class_type_click_count == 0: #
            self.controller.sortProcessed("Class_Type") #
            self.__reset_click_count()
            self.class_type_click_count += 1 #
        elif self.class_type_click_count == 1: #
            self.controller.sortProcessed("Class_Type", descending=True) #
            self.__reset_click_count()
        self.__loadSchedules()
        self.update()
    
    def __location_clicked(self):
        self.__clearSchedulesTable()
        if self.location_click_count == 0: #
            self.controller.sortProcessed("Location") #
            self.__reset_click_count()
            self.location_click_count += 1 #
        elif self.location_click_count == 1: #
            self.controller.sortProcessed("Location", descending=True) #
            self.__reset_click_count()
        self.__loadSchedules()
        self.update()
    
    def __size_clicked(self):
        self.__clearSchedulesTable()
        if self.size_click_count == 0: #
            self.controller.sortProcessed("Size") #
            self.__reset_click_count()
            self.size_click_count += 1 #
        elif self.size_click_count == 1: #
            self.controller.sortProcessed("Size", descending=True) #
            self.__reset_click_count()
        self.__loadSchedules()
        self.update()
    
    def __zone_clicked(self):
        self.__clearSchedulesTable()
        if self.zone_click_count == 0: #
            self.controller.sortProcessed("Zone") #
            self.__reset_click_count()
            self.zone_click_count += 1 #
        elif self.zone_click_count == 1: #
            self.controller.sortProcessed("Zone", descending=True) #
            self.__reset_click_count()
        self.__loadSchedules()
        self.update()

    # date & range of dates
    def __boxChecked(self):
        if (self.checked.get() == 1): # when range
            # destroy single
            self.date_label.destroy()
            self.date_space_label.destroy()
            self.date_option.destroy()

            self.start_date_label = customtkinter.CTkLabel(self.tabview.tab("Date & Time"), text="Start Date")
            self.start_date_label.grid(row=0, column=4, sticky='e')
            
            self.end_date_label = customtkinter.CTkLabel(self.tabview.tab("Date & Time"), text="End Date")
            self.end_date_label.grid(row=0, column=6, sticky='w')

            self.start_date_option = customtkinter.CTkEntry(self.tabview.tab("Date & Time"), height=15, width=95) # xxx
            self.start_date_option.insert(0, "dd/mm/yyyy")
            self.start_date_option.grid(row=1, column=4, sticky='e')
            self.start_date_option.bind("<1>", self.__start_date_option_clicked)

            self.range_space_label = customtkinter.CTkLabel(self.tabview.tab("Date & Time"), text="~")
            self.range_space_label.grid(row=1, column=5, padx=5)

            self.end_date_option = customtkinter.CTkEntry(self.tabview.tab("Date & Time"), height=15, width=95) # xxx
            self.end_date_option.insert(0, "dd/mm/yyyy")
            self.end_date_option.grid(row=1, column=6, sticky="w")
            self.end_date_option.bind("<1>", self.__end_date_option_clicked)

        else: # when single
            # destroy double
            self.start_date_label.destroy()
            self.end_date_label.destroy()
            self.start_date_option.destroy()
            self.range_space_label.destroy()
            self.end_date_option.destroy()

            self.date_label = customtkinter.CTkLabel(self.tabview.tab("Date & Time"), text="Date")
            self.date_label.grid(row=0, column=4, pady=(10,0), sticky='e')
            self.date_space_label = customtkinter.CTkLabel(self.tabview.tab("Date & Time"), text=":")
            self.date_space_label.grid(row=0, column=5, padx=5, pady=(10,0))
            self.date_option = customtkinter.CTkEntry(self.tabview.tab("Date & Time"), height=15, width=95) # xxx
            self.date_option.insert(0, "dd/mm/yyyy")
            self.date_option.grid(row=0, column=6, pady=(10,0), sticky="w")
            self.date_option.bind("<1>", self.__date_option_clicked)

    def __date_option_clicked(self, event):
        self.date_window = customtkinter.CTkToplevel()
        self.date_window.grab_set()
        self.date_window.title("Select Schedule Date")
        self.date_window.geometry("250x220")
        self.date_window.grid_rowconfigure((0,1), weight=1)
        self.date_window.grid_columnconfigure(0, weight=1)
        self.cal = Calendar(self.date_window, selectmode="day", date_pattern="dd/mm/y")
        self.cal.grid(row=0, column=0)
        self.submit_button = customtkinter.CTkButton(self.date_window, text="Submit", command=self.__date_picked)
        self.submit_button.grid(row=1, column=0)

    def __date_picked(self):
        self.date_option.delete(0, tkinter.END)
        self.date_option.insert(0, self.cal.get_date())
        self.date_window.destroy()

    def __start_date_option_clicked(self, event):
        max_date = self.end_date_option.get()
        self.date_window = customtkinter.CTkToplevel()
        self.date_window.grab_set()
        self.date_window.title("Select Schedule Start Date")
        self.date_window.geometry("250x220")
        self.date_window.grid_rowconfigure((0,1), weight=1)
        self.date_window.grid_columnconfigure(0, weight=1)
        if max_date != "dd/mm/yyyy":
            self.cal = Calendar(self.date_window, selectmode="day", date_pattern="dd/mm/y", maxdate=datetime.strptime(max_date, "%d/%m/%Y"))
        else:
            self.cal = Calendar(self.date_window, selectmode="day", date_pattern="dd/mm/y")
        self.cal.grid(row=0, column=0)
        self.submit_button = customtkinter.CTkButton(self.date_window, text="Submit", command=self.__start_date_picked)
        self.submit_button.grid(row=1, column=0)

    def __start_date_picked(self):
        self.start_date_option.delete(0, tkinter.END)
        self.start_date_option.insert(0, self.cal.get_date())
        self.date_window.destroy()

    def __end_date_option_clicked(self, event):
        min_date = self.start_date_option.get()
        self.date_window = customtkinter.CTkToplevel()
        self.date_window.grab_set()
        self.date_window.title("Select Schedule End Date")
        self.date_window.geometry("250x220")
        self.date_window.grid_rowconfigure((0,1), weight=1)
        self.date_window.grid_columnconfigure(0, weight=1)
        if min_date != "dd/mm/yyyy":
            self.cal = Calendar(self.date_window, selectmode="day", date_pattern="dd/mm/y", mindate=datetime.strptime(min_date, "%d/%m/%Y"))
        else:
            self.cal = Calendar(self.date_window, selectmode="day", date_pattern="dd/mm/y")
        self.cal.grid(row=0, column=0)
        self.submit_button = customtkinter.CTkButton(self.date_window, text="Submit", command=self.__end_date_picked)
        self.submit_button.grid(row=1, column=0)

    def __end_date_picked(self):
        self.end_date_option.delete(0, tkinter.END)
        self.end_date_option.insert(0, self.cal.get_date())
        self.date_window.destroy()
    
    # popup
    def errorPopup(self, message):
        messagebox.showerror("An error occurred", message)
        self.update()

    # events
    def importButtonPressed(self):
        self.__clearFilesTable()
        self.file_path = customtkinter.filedialog.askdirectory() + "/"
        self.entry.configure(state="normal")
        self.entry.delete(0, "end")
        self.entry.insert(0,f"{self.file_path}")
        self.entry.configure(state="disabled")
        self.reloadButtonPressed()
        self.__loadFiles()
        if len(self.handler.getFiles()) == 0:
            self.errorPopup("No valid CSV file found!")
        self.update()

    def removeButtonPressed(self):
        # class_type = [self.class_type_option.get(i) for i in self.class_type_option.curselection()]
        self.selected_item = [i for i in self.files_table.selection()]
        for item in self.selected_item:
            self.handler.addIgnore(self.files_table.item(item)["values"][1])
            self.files_table.delete(item)
            
        self.reloadButtonPressed()

    def exportButtonPressed(self):
        if len(self.controller.getProcessed()) == 0:
            self.errorPopup("No Schedules Found!")
        else:
            self.confirmButtonPressed()
            self.export_window = customtkinter.CTkToplevel()
            self.export_window.grab_set()
            self.export_window.title("Export Options")
            self.export_window.geometry("325x130")
            self.export_window.grid_rowconfigure((0,1,2), weight=1)
            self.export_window.grid_columnconfigure((0,1,2), weight=1)
            self.export_window.grid_columnconfigure(3, weight=0)
            
            # path entry
            self.export_path_entry = customtkinter.CTkEntry(self.export_window, placeholder_text="Set Path")
            self.export_path_entry.insert(0, "Set Path")
            self.export_path_entry.configure(state="disabled")
            self.export_path_entry.grid(row=0, column=0, columnspan=3, padx=(5,0), pady=(5,0), sticky="we")

            # open button
            self.open_button = customtkinter.CTkButton(self.export_window, text="Open", width=70, font=customtkinter.CTkFont(size=12, weight="bold"), fg_color="transparent", border_width=2, text_color=("gray10", "#DCE4EE"), command=self.openButtonPressed)
            self.open_button.grid(row=0, column=3, sticky='we', padx=5, pady=(5,0))

            # name entry
            self.file_name_entry = customtkinter.CTkEntry(self.export_window, placeholder_text="File Name")
            self.file_name_entry.grid(row=1, column=0, columnspan=3, padx=(5,0), pady=(5,0), sticky="we")

            # format option
            self.format_set = ["pdf", "xlsx"]
            self.format_option = customtkinter.CTkComboBox(self.export_window, values=self.format_set, width=70)
            self.format_option.set("pdf")
            self.format_option.grid(row=1, column=3, sticky='we', padx=5, pady=(5,0))

            self.save_button = customtkinter.CTkButton(self.export_window, width=70, text="Save", font=customtkinter.CTkFont(size=12, weight="bold"), border_width=2, border_color="grey", text_color=("gray10", "#DCE4EE"), command=self.saveButtonPressed)
            self.save_button.grid(row=2, column=3, sticky='we', padx=5, pady=(0,10))

    def openButtonPressed(self):
        self.export_path = customtkinter.filedialog.askdirectory() + "/"
        self.export_path_entry.configure(state="normal")
        self.export_path_entry.delete(0, "end")
        self.export_path_entry.insert(0,f"{self.export_path}")
        self.export_path_entry.configure(state="disabled")

    def saveButtonPressed(self):
        export_path = self.export_path_entry.get()
        file_name = self.file_name_entry.get()
        format = self.format_option.get()
        self.builder = TimetableBuilder(self.controller)
        self.builder.export(format,file_name,export_path)
        self.export_window.destroy()
        self.confirmButtonPressed()
        self.errorPopup("Successfully Saved")

    def reloadButtonPressed(self):
        self.handler.resetHandler()
        self.handler.loadDirectory(self.file_path)
        self.controller = ScheduleController(self.handler)
        self.__clearSchedulesTable()
        self.__loadSchedules()
        self.__reload_filters()
        self.update()

    def confirmButtonPressed(self, called=False):
        self.handler.resetHandler()
        self.handler.loadDirectory(self.file_path)
        self.controller = ScheduleController(self.handler)
        self.__clearSchedulesTable()


        cohort = self.cohort_option.get()
        if cohort == "No Filter":
            cohort = None

        study_mode = self.study_mode_option.get()
        if study_mode == "No Filter":
            study_mode = None

        lecturer = self.lecturer_option.get()
        if lecturer == "No Filter":
            lecturer = None

        module_code = self.module_code_option.get()
        if module_code == "No Filter":
            module_code = None

        if self.checked.get() == 0:
            start_date = None
            end_date = None
            date = self.date_option.get()
            if date == "dd/mm/yyyy":
                date = None
            else:
                date = datetime.strptime(date, '%d/%m/%Y')

        else:
            date = None
            start_date = self.start_date_option.get()
            if start_date == "dd/mm/yyyy":
                start_date = None
            else:
                start_date = datetime.strptime(start_date, '%d/%m/%Y')

            end_date = self.end_date_option.get()
            if end_date == "dd/mm/yyyy":
                end_date = None
            else:
                end_date = datetime.strptime(end_date, '%d/%m/%Y')

        duration = self.duration_option.get()
        if duration == "No Filter":
            duration = None

        start_time = self.start_time_option.get()
        if start_time == "No Filter":
            start_time = None
        else:
            start_time = datetime.strptime(start_time, '%H:%M:%S')

        end_time = self.end_time_option.get()
        if end_time == "No Filter":
            end_time = None
        else:
            end_time = datetime.strptime(end_time, '%H:%M:%S')

        location = self.location_option.get()
        if location == "No Filter":
            location = None

        size = self.size_option.get()
        if size == "No Filter":
            size = None
        else:
            size = int(size)

        zone = self.zone_option.get()
        if zone == "No Filter":
            zone = None

        description = [self.description_option.get(i) for i in self.description_option.curselection()]
        if description == []:
            description = None
        else:
            description = "&&&".join(description)

        day = [self.day_option.get(i) for i in self.day_option.curselection()]
        if day == []:
            day = None
        else:
            day = "&&&".join(day)

        class_type = [self.class_type_option.get(i) for i in self.class_type_option.curselection()]
        if class_type == []:
            class_type = None
        else:
            class_type = "&&&".join(class_type)

        self.controller.control(sortBy="Date", Cohort=cohort, Study_Mode=study_mode, Lecturer=lecturer, Module_Code=module_code, Date=date, Start_Date=start_date, End_Date=end_date, Duration=duration, Start_Time=start_time, End_Time=end_time, Location=location, Size=size, Zone=zone, Description=description, Day_str=day, Class_Type=class_type)


        self.__loadSchedules()
        self.update()

if __name__ == "__main__":
    app = App()
    app.mainloop()